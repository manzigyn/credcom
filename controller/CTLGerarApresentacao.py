import pandas as pd
from controller import CTLParametrizacao
from controller import CTLLogGeracao
from utils import utilidades as ut
from entity import ApresentacaoExcel as enAp
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import openpyxl
import os
from datetime import datetime
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
#import xlwings as xw
#from pyxll import xl_func, plot
#import plotly.express as px

class CTLApresentacaoExcel:
    
    def __init__(self, entidade: enAp.ApresentacaoExcel):
        self.parametrizacao = CTLParametrizacao.carregar()
        self.apresentacaoExcel = entidade
    
    def haCarga(self) -> bool:
        return self.parametrizacao.haCarga

    def obterTemplate(self, pasta_origem,  chave):
        for arquivo in os.listdir(pasta_origem):
            arquivo = str(arquivo).lower()
            if arquivo.endswith(".xlsx"): 
                if any(x in arquivo for x in [f'template_{chave}.xlsx'.lower()]):
                    return arquivo
        return ""
                
                
    def gerarPeloTemplate(self)-> openpyxl.Workbook:
        try:

            arquivo_origem = f'{self.parametrizacao.dir_template}template.xlsx'
            self.apresentacaoExcel.nomeArquivo = f'{self.apresentacaoExcel.contratante.nome}_{self.apresentacaoExcel.contratante.ano}_{self.apresentacaoExcel.contratante.mes}_{datetime.today().strftime('%Y%m%d')}_{datetime.today().strftime('%H%M%S')}.xlsx'
            arquivo_destino = f'{self.parametrizacao.dir_apresentacao}{self.apresentacaoExcel.nomeArquivo}'
            ut.copiarArquivo(arquivo_origem, arquivo_destino )
            
            workbook =  load_workbook(arquivo_destino)
            worksheet= workbook['Apresentação']
            
            worksheet['B2'] = f'Apresentação de Resultados - {self.apresentacaoExcel.contratante.nome}'
            worksheet['C7'] = self.apresentacaoExcel.sms
            worksheet['C8'] = self.apresentacaoExcel.ligacao
            
            linha =11
            
            #loteamento valor caixa
            linha_formula = linha+len(self.apresentacaoExcel.df_loteamento_caixa.index)
            formula=f'SOMA(C{linha+1}:C{linha_formula-1})'            
            soma = self.apresentacaoExcel.df_loteamento_caixa["Total"].sum()
            self.write_df_excel(self.apresentacaoExcel.df_loteamento_caixa[["Loteamento","Total"]].rename(columns={"Total":"Valor em Caixa"}), worksheet, linha,2, soma, linha_formula,3)
            linha_lot_val1 = linha
                        
            #loteamento valor recuperado
            linha_formula = linha+len(self.apresentacaoExcel.df_loteamento_recuperado.index)            
            formula=f'SOMA(F{linha+1}:F{linha_formula-1})'            
            soma = self.apresentacaoExcel.df_loteamento_recuperado["Total"].sum()
            self.write_df_excel(self.apresentacaoExcel.df_loteamento_recuperado[["Loteamento","Total"]].rename(columns={"Total":"Valor Recuperado"}), worksheet, linha,5, soma, linha_formula,6)
            
            #loteamento tipo
            linha_formula = linha+len(self.apresentacaoExcel.df_loteamento_tipo.index)
            formula=f'SOMA(J{linha+1}:J{linha_formula-1})'            
            soma = self.apresentacaoExcel.df_loteamento_tipo["Total"].sum()
            self.write_df_excel(self.apresentacaoExcel.df_loteamento_tipo[["Tipo","Qtde","Total"]].rename(columns={"Total":"Valor Recuperado"}), worksheet, linha,8, soma, linha_formula,10)

            linha += len(self.apresentacaoExcel.df_loteamento_caixa.index)
            if linha < len(self.apresentacaoExcel.df_loteamento_recuperado.index):
                linha = len(self.apresentacaoExcel.df_loteamento_recuperado.index)
            
            
            linha += 4    
            
            #loteamento status virtua
            linha_formula = linha+len(self.apresentacaoExcel.df_loteamento_statusvirtua.index)
            formula=f'SOMA(D{linha+1}:D{linha_formula-1})'            
            soma = self.apresentacaoExcel.df_loteamento_statusvirtua["TotalInadimplencia"].sum()
            self.write_df_excel(self.apresentacaoExcel.df_loteamento_statusvirtua[["Loteamento","Qtde","TotalInadimplencia","% Recuperada"]].rename(columns={"TotalInadimplencia":"Valor de Inadimplência"}), worksheet, linha,2, soma, linha_formula,4)
            linha += len(self.apresentacaoExcel.df_loteamento_statusvirtua.index) + 4
            
            pie = PieChart()
            labels = Reference(worksheet, min_col=2, min_row=linha_lot_val1, max_row=len(self.apresentacaoExcel.df_loteamento_caixa.index))
            data = Reference(worksheet, min_col=3, min_row=linha_lot_val1, max_row=len(self.apresentacaoExcel.df_loteamento_caixa.index))
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Loteamento por Valor em Caixa"
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True

            # Cut the first slice out of the pie
            slice = DataPoint(idx=0, explosion=20)
            pie.series[0].data_points = [slice]

            worksheet.add_chart(pie, f"D{linha}")
            
            #img_bytes  = self.fig_loteamento_caixa.to_image(format="png")
            #img = openpyxl.drawing.image.Image(img_bytes)
            #worksheet.add_image(img,f'B{row}')
            
            #status virtua quantidade
            self.write_df_excel(self.apresentacaoExcel.df_statusvirtua_quantidade, worksheet, linha,2)
            #row += len(self.apresentacaoExcel.df_loteamento_statusvirtua.index) + 4
            #https://docs.xlwings.org/en/stable/matplotlib.html
            #worksheet.pictures.add(self.apresentacaoExcel.gr_loteamento_caixa, name='Loteamento Caixa', update=True)
            
            workbook.save(arquivo_destino)
            self.apresentacaoExcel.caminhoArquivo = arquivo_destino
            CTLLogGeracao.salvar(self.apresentacaoExcel.contratante, self.apresentacaoExcel.nomeArquivo, self.apresentacaoExcel.caminhoArquivo)
            return workbook
        except PermissionError as e:
            e.add_note(f'Não foi possível gerar o arquivo {arquivo_destino}, pois deve estar aberto ')
            raise
        except Exception as e:
            e.add_note(f'Não foi possível gerar o arquivo {arquivo_destino}, devido ao erro {e}')
            raise
    
    def write_df_excel(self, df, ws, startrow=0, startcol=0, formula=0, linha_formula=-1, coluna_formula=-1, Header=True, Index=False):
        """Write DataFrame df to openpyxl worksheet ws"""
        startrow-=1
        startcol-=1
        rows = dataframe_to_rows(df, header=Header, index=Index)

       
        for r_idx, row in enumerate(rows, startrow + 1):
            for c_idx, value in enumerate(row, startcol + 1):
                ws.cell(row=r_idx, column=c_idx).value = value
                if isinstance(value, float):
                    ws.cell(row=r_idx, column=c_idx).number_format ='R$ #,##0.00'
                if r_idx == startrow + 1: #Cabeçalho
                    ws.cell(row=r_idx, column=c_idx).font = self.fonteCabecalhoDF()
                    ws.cell(row=r_idx, column=c_idx).fill = self.fundoCabecalhoDF()
                else:
                    if c_idx > startcol + 1:
                        ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="right")
        if linha_formula > 0 and coluna_formula > 0:
            ws.cell(row=linha_formula, column=coluna_formula).value = formula #f'={formula}'
            #ws.cell(row=linha_formula, column=coluna_formula).value = ws.cell(row=linha_formula, column=coluna_formula).value.replace("==","=")

                
    def fonteCabecalhoDF(self):
        return Font(name='Calibri',
                 size=12,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='00FFFFFF')
    
    def fundoCabecalhoDF(self):
        return PatternFill(fill_type="solid",
                 start_color='00003366',
                 end_color='00003366')
        
    #https://stackoverflow.com/questions/65272408/plotly-how-to-embed-a-fully-interactive-plotly-figure-in-excel
    #@xl_func
    #def plotly_plot(fig):
    #    plot(fig)