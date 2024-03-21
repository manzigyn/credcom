import pandas as pd
from controller import CTLParametrizacao
from utils import utilidades as ut
from entity import ApresentacaoExcel as enAp
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import openpyxl
import os


class CTLApresentacaoExcel:
    
    def __init__(self, entidade: enAp.ApresentacaoExcel):
        self.parametrizacao = CTLParametrizacao.carregar()
        self.apresentacaoExcel = entidade
    

    def obterTemplate(self, pasta_origem,  chave):
        for arquivo in os.listdir(pasta_origem):
            arquivo = str(arquivo).lower()
            if arquivo.endswith(".xlsx"): 
                if any(x in arquivo for x in [f'template_{chave}.xlsx'.lower()]):
                    return arquivo
        return ""
                
                
    def gerarPeloTemplate(self)-> openpyxl.Workbook:
        try:

            arquivo_origem = f'{self.parametrizacao.dir_template}template.xlsx'
            self.apresentacaoExcel.NomeArquivo = f'{self.apresentacaoExcel.Contratante}_{self.apresentacaoExcel.Ano}_{self.apresentacaoExcel.Mes}.xlsx'
            arquivo_destino = f'{self.parametrizacao.dir_apresentacao}{self.apresentacaoExcel.NomeArquivo}'
            ut.copiarArquivo(arquivo_origem, arquivo_destino )
            
            workbook =  load_workbook(arquivo_destino)
            worksheet= workbook['Apresentação']
            
            worksheet['B2'] = f'Apresentação de Resultados - {self.apresentacaoExcel.Contratante}'
            worksheet['C7'] = self.apresentacaoExcel.Sms
            worksheet['C8'] = self.apresentacaoExcel.Ligacao
            
            row =11
            linha_formula = row+len(self.apresentacaoExcel.df_loteamento_caixa.index)
            formula=f'SOMA(C{row+1}:C{linha_formula-1})'            
            self.write_df_excel(self.apresentacaoExcel.df_loteamento_caixa[["Loteamento","Total"]].rename(columns={"Total":"Valor em Caixa"}), worksheet, row,2, formula, linha_formula,3)
            
            linha_formula = row+len(self.apresentacaoExcel.df_loteamento_recuperado.index)
            formula=f'SOMA(F{row+1}:F{linha_formula-1})'            
            self.write_df_excel(self.apresentacaoExcel.df_loteamento_recuperado[["Loteamento","Total"]].rename(columns={"Total":"Valor Recuperado"}), worksheet, row,5,formula, linha_formula,6)
            
            linha_formula = row+len(self.apresentacaoExcel.df_loteamento_tipo.index)
            formula=f'SOMA(J{row+1}:J{linha_formula-1})'            
            self.write_df_excel(self.apresentacaoExcel.df_loteamento_tipo[["Tipo","Qtde","Total"]].rename(columns={"Total":"Valor Recuperado"}), worksheet, row,8,formula, linha_formula,10)

            row += len(self.apresentacaoExcel.df_loteamento_caixa.index)
            if row < len(self.apresentacaoExcel.df_loteamento_recuperado.index):
                row = len(self.apresentacaoExcel.df_loteamento_recuperado.index)
            
            row += 4    
            linha_formula = row+len(self.apresentacaoExcel.df_loteamento_statusvirtua.index)
            formula=f'SOMA(D{row+1}:D{linha_formula-1})'            

            self.write_df_excel(self.apresentacaoExcel.df_loteamento_statusvirtua[["Loteamento","Qtde","TotalInadimplencia","% Recuperada"]].rename(columns={"TotalInadimplencia":"Valor de Inadimplência"}), worksheet, row,2, formula, linha_formula,4)
            row += len(self.apresentacaoExcel.df_loteamento_statusvirtua.index) + 4
            
            
            #img_bytes  = self.fig_loteamento_caixa.to_image(format="png")
            #img = openpyxl.drawing.image.Image(img_bytes)
            #worksheet.add_image(img,f'B{row}')
            
            self.write_df_excel(self.apresentacaoExcel.df_statusvirtua_quantidade, worksheet, row,2)
            
            workbook.save(arquivo_destino)
            self.apresentacaoExcel.CaminhoArquivo = arquivo_destino
            return workbook
        except PermissionError as e:
            e.add_note(f'Não foi possível gerar o arquivo {arquivo_destino}, pois deve estar aberto ')
            raise
        except Exception as e:
            e.add_note(f'Não foi possível gerar o arquivo {arquivo_destino}, devido ao erro {e}')
            raise
    
    def write_df_excel(self, df, ws, startrow=0, startcol=0, formula='', linha_formula=-1, coluna_formula=-1, Header=True, Index=False):
        """Write DataFrame df to openpyxl worksheet ws"""
        startrow-=1
        startcol-=1
        rows = dataframe_to_rows(df, header=Header, index=Index)

       
        for r_idx, row in enumerate(rows, startrow + 1):
            for c_idx, value in enumerate(row, startcol + 1):
                ws.cell(row=r_idx, column=c_idx).value = value
                if isinstance(value, float):
                    ws.cell(row=r_idx, column=c_idx).number_format ='R$ #,##0.00'
                if r_idx == startrow + 1:
                    ws.cell(row=r_idx, column=c_idx).font = self.fonteCabecalhoDF()
                    ws.cell(row=r_idx, column=c_idx).fill = self.fundoCabecalhoDF()
                else:
                    if c_idx > startcol + 1:
                        ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="right")
        if linha_formula > 0 and coluna_formula > 0:
            ws.cell(row=linha_formula, column=coluna_formula).value = f'={formula}'
            ws.cell(row=linha_formula, column=coluna_formula).value = ws.cell(row=linha_formula, column=coluna_formula).value.replace("==","=")

                
    def fonteCabecalhoDF(self):
        return Font(name='Calibri',
                 size=12,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='00FFFFFF')
    
    def fundoCabecalhoDF(self):
        return PatternFill(fill_type="solid",
                 start_color='00003366',
                 end_color='00003366')