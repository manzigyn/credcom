from dataclasses import dataclass
from matplotlib.figure import Figure
import pandas as pd
from utils import utilidades as ut
import plotly_express as pl
import os
from utils import pandas_files as pf
from model import DBPagamento as dbPgto
from model import DBDistribuicao as dbDis
from model import DBConfiguracaoResultado as dbConf
from classes import Entidades as ent
from io import BytesIO
import xlsxwriter as xls
from os.path import expanduser
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import openpyxl
import math


    
@dataclass
class Apresentacao():
        df_loteamento_caixa: pd.DataFrame
        df_loteamento_recuperado: pd.DataFrame
        df_loteamento_tipo: pd.DataFrame
        df_loteamento_statusvirtua: pd.DataFrame
        df_statusvirtua: pd.DataFrame
        gr_loteamento_caixa: pl.pie
        gr_loteamento_recuperado: pl.pie
        gr_statusvirtua: pl.bar
        
@dataclass
class ExcelDataFrameGraphis():
    def gerar_dfs_graphs(self, df_list, gr_list, sheets, file_name, spaces):
        writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
        row = 0
        index=0
        workbook=writer.book
        worksheet=workbook.add_worksheet(sheets)
        #worksheet = writer.sheets[sheets]
        for dataframe in df_list:
            dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
            #if index < len(gr_list):                
            #    image_data = BytesIO(gr_list[index].to_image(format="png"))
            #    worksheet.insert_image(row, 5, f'graph{index}.png', {'image_data': image_data})
            #    index+=1
            row = row + len(dataframe.index) + spaces + 1
        writer.close()
        return writer
        
    def gerar_dfs(self, df_list, sheets, file_name, spaces):
        writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
        row = 0
        for dataframe in df_list:
            dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
            row = row + len(dataframe.index) + spaces + 1
        writer.save()

    def gerar_dfs_tabs(df_list, sheet_list, file_name):
        writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
        for dataframe, sheet in zip(df_list, sheet_list):
            dataframe.to_excel(writer, sheet_name=sheet, startrow=0 , startcol=0)   
        writer.save()

class ApresentacaoExcel:
    def __init__(self, 
                 sContratante: str,
                 iAno: int,
                 iMes: int,
                 df_loteamento_caixa: pd.DataFrame, 
                 df_loteamento_recuperado: pd.DataFrame, 
                 df_loteamento_tipo: pd.DataFrame,
                 df_loteamento_statusvirtua: pd.DataFrame,
                 df_statusvirtua_quantidade: pd.DataFrame,
                 iSms: int,
                 iLigacao: int, 
                 fig_loteamento_caixa: Figure, 
                 sArquivoGeradoNome = '',
                 sArquivoGeradoCaminho = ''):
        self.sContratante = sContratante
        self.iAno = iAno
        self.iMes = iMes
        self.df_loteamento_caixa = df_loteamento_caixa
        self.df_loteamento_recuperado = df_loteamento_recuperado
        self.df_loteamento_tipo = df_loteamento_tipo
        self.df_loteamento_statusvirtua = df_loteamento_statusvirtua
        self.df_statusvirtua_quantidade = df_statusvirtua_quantidade
        self.iSms = iSms
        self.iLigacao =iLigacao
        self.fig_loteamento_caixa = fig_loteamento_caixa
        self.sArquivoGeradoNome = sArquivoGeradoNome
        self.sArquivoGeradoCaminho = sArquivoGeradoCaminho
    
    
    def gerarArquivo(self) -> pd.ExcelWriter:
        nomeArquivo = f'dados/apresentacao/{self.sContratante}_{self.iAno}_{self.iMes}.xlsx'
        guia = 'Apresentacao'
        
        writer = pd.ExcelWriter(nomeArquivo,engine='xlsxwriter')   
        workbook=writer.book
        worksheet=workbook.add_worksheet(guia)
        
        cell_format = workbook.add_format({'bold': True, 'font_color': 'gray', 'font_size': 36})
        worksheet.set_column("B:J", 12)
        worksheet.set_row(3, 30)
    
        merge_format = workbook.add_format(
                {
                    "bold": 1,
                    "border": 0,
                    "align": "center",
                    "valign": "vcenter",
                    'font_color': 'gray',
                    'font_size': 24
                }
            )
        worksheet.merge_range("B2:J2", f'Apresentação de Resultados - {self.sContratante} Periodo {self.iMes}/{self.iAno}', merge_format)
        #worksheet.write(1,1, f'Apresentação de Resultados - {self.sContratante} Periodo {self.iMes}/{self.iAno}',cell_format)
         
        border_fmt = workbook.add_format({'bottom':5, 'top':5, 'left':5, 'right':5})
        sizeLoteamento = 50
        worksheet.set_column('A:A', 2)
        worksheet.set_column('B:B', sizeLoteamento)
        worksheet.set_column('C:C', 15)
        
        worksheet.set_column('D:D', 5)
        
        worksheet.set_column('E:E', sizeLoteamento)
        worksheet.set_column('F:F', 15)
        row = 5
        self.df_loteamento_caixa.to_excel(writer,sheet_name=guia,startrow=row , startcol=1,index=False,merge_cells=True) 
        self.df_loteamento_recuperado.to_excel(writer,sheet_name=guia,startrow=row , startcol=4,index=False) 
        self.df_loteamento_tipo.to_excel(writer,sheet_name=guia,startrow=row , startcol=7,index=False) 
        
        row += len(self.df_loteamento_caixa.index)
        if row < len(self.df_loteamento_recuperado.index):
            row = len(self.df_loteamento_recuperado.index)
        row += 4    
        self.df_loteamento_statusvirtua.to_excel(writer,sheet_name=guia,startrow=row , startcol=1, index=False) 
        
        row += len(self.df_loteamento_statusvirtua.index) + 4
        self.df_statusvirtua_quantidade.to_excel(writer,sheet_name=guia,startrow=row , startcol=1, index=False) 

        writer.close()
        return writer
    
    
    def obterTemplate(self, pasta_origem,  chave):
        for arquivo in os.listdir(pasta_origem):
            arquivo = str(arquivo).lower()
            if arquivo.endswith(".xlsx"): 
                if any(x in arquivo for x in [f'template_{chave}.xlsx'.lower()]):
                    return arquivo
        return ""
                
                
    def gerarPeloTemplate(self)-> openpyxl.Workbook:
        try:
            pasta = 'dados/'
            origem = f'{pasta}template/'
            #template = self.obterTemplate(origem, self.sContratante)            
            #if not template:
            #    return
            template = 'template.xlsx'
            arquivo_origem = f'{origem}{template}'
            self.sArquivoGeradoNome = f'{self.sContratante}_{self.iAno}_{self.iMes}.xlsx'
            arquivo_destino = f'D:/pessoal/Magalhaes/{self.sArquivoGeradoNome}'
            ut.copiarArquivo(arquivo_origem, arquivo_destino )
            
            workbook =  load_workbook(arquivo_destino)
            worksheet= workbook['Apresentação']
            
            worksheet['B2'] = f'Apresentação de Resultados - {self.sContratante}'
            worksheet['C7'] = self.iSms
            worksheet['C8'] = self.iLigacao
            
            row =11
            linha_formula = row+len(self.df_loteamento_caixa.index)
            formula=f'SOMA(C{row+1}:C{linha_formula-1})'            
            self.write_df_excel(self.df_loteamento_caixa[["Loteamento","Total"]].rename(columns={"Total":"Valor em Caixa"}), worksheet, row,2, formula, linha_formula,3)
            
            linha_formula = row+len(self.df_loteamento_recuperado.index)
            formula=f'SOMA(F{row+1}:F{linha_formula-1})'            
            self.write_df_excel(self.df_loteamento_recuperado[["Loteamento","Total"]].rename(columns={"Total":"Valor Recuperado"}), worksheet, row,5,formula, linha_formula,6)
            
            linha_formula = row+len(self.df_loteamento_tipo.index)
            formula=f'SOMA(J{row+1}:J{linha_formula-1})'            
            self.write_df_excel(self.df_loteamento_tipo[["Tipo","Qtde","Total"]].rename(columns={"Total":"Valor Recuperado"}), worksheet, row,8,formula, linha_formula,10)

            row += len(self.df_loteamento_caixa.index)
            if row < len(self.df_loteamento_recuperado.index):
                row = len(self.df_loteamento_recuperado.index)
            
            row += 4    
            linha_formula = row+len(self.df_loteamento_statusvirtua.index)
            formula=f'SOMA(D{row+1}:D{linha_formula-1})'            

            self.write_df_excel(self.df_loteamento_statusvirtua[["Loteamento","Qtde","TotalInadimplencia","% Recuperada"]].rename(columns={"TotalInadimplencia":"Valor de Inadimplência"}), worksheet, row,2, formula, linha_formula,4)
            row += len(self.df_loteamento_statusvirtua.index) + 4
            
            
            #img_bytes  = self.fig_loteamento_caixa.to_image(format="png")
            #img = openpyxl.drawing.image.Image(img_bytes)
            #worksheet.add_image(img,f'B{row}')
            
            self.write_df_excel(self.df_statusvirtua_quantidade, worksheet, row,2)
            
            workbook.save(arquivo_destino)
            self.sArquivoGeradoCaminho = arquivo_destino
            return workbook
        except PermissionError as e:
            e.add_note(f'Não foi possível gerar o arquivo {arquivo_destino}, pois deve estar aberto ')
            raise
        except Exception as e:
            e.add_note(f'Não foi possível gerar o arquivo {arquivo_destino}, devido ao erro {e}')
            raise
    
    def write_df_excel(self, df, ws, startrow=0, startcol=0, formula='', linha_formula=-1, coluna_formula=-1, Header=True, Index=False):
        """Write DataFrame df to openpyxl worksheet ws"""
        startrow-=1
        startcol-=1
        rows = dataframe_to_rows(df, header=Header, index=Index)

       
        for r_idx, row in enumerate(rows, startrow + 1):
            for c_idx, value in enumerate(row, startcol + 1):
                ws.cell(row=r_idx, column=c_idx).value = value
                if isinstance(value, float):
                    ws.cell(row=r_idx, column=c_idx).number_format ='R$ #,##0.00'
                if r_idx == startrow + 1:
                    ws.cell(row=r_idx, column=c_idx).font = self.fonteCabecalhoDF()
                    ws.cell(row=r_idx, column=c_idx).fill = self.fundoCabecalhoDF()
                else:
                    if c_idx > startcol + 1:
                        ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="right")
        if linha_formula > 0 and coluna_formula > 0:
            ws.cell(row=linha_formula, column=coluna_formula).value = f'={formula}'
            ws.cell(row=linha_formula, column=coluna_formula).value = ws.cell(row=linha_formula, column=coluna_formula).value.replace("==","=")

                
    def fonteCabecalhoDF(self):
        return Font(name='Calibri',
                 size=12,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='00FFFFFF')
    
    def fundoCabecalhoDF(self):
        return PatternFill(fill_type="solid",
                 start_color='00003366',
                 end_color='00003366')